#!/usr/bin/env python3
"""Build an Excel spreadsheet (.xlsx) from report payload.

Creates a multi-tab workbook with:
  Tab 1: Summary — key metrics
  Tab 2: Channels — per-channel breakdown
  Tab 3: Time Series — raw time-bucketed data
  Tab 4: Platform Split — platform breakdown per day
  Tab 5: Language Split — language breakdown per day

Input JSON (via stdin):
{
  "payload": { ... full report payload ... },
  "timeSeries": [ { "timestamp": "...", "totalCCV": 123, "channelCount": 3 }, ... ],
  "outputPath": "/tmp/reports/{id}/report.xlsx",
  "branding": { ... }
}
"""

import sys
import os
import json
from typing import Any, Dict, List, Optional

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from report_helpers import (
    read_stdin_json, load_branding, hex_to_rgb,
    platform_label, tier_label,
    compact_number, format_number, format_hours, format_date, format_datetime,
    aggregate_metrics, build_scope_label,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Styling ──────────────────────────────────────────────────────────────────

def header_fill(branding: Dict) -> PatternFill:
    r, g, b = branding.get('table_header_bg_rgb', (30, 41, 59))
    return PatternFill(start_color=f'{r:02X}{g:02X}{b:02X}', end_color=f'{r:02X}{g:02X}{b:02X}', fill_type='solid')


def header_font() -> Font:
    return Font(name='Arial', size=10, bold=True, color='FFFFFF')


def data_font() -> Font:
    return Font(name='Arial', size=10, color='374151')


def number_font() -> Font:
    return Font(name='Arial', size=10, color='374151')


def alt_fill() -> PatternFill:
    return PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')


def thin_border() -> Border:
    side = Side(style='thin', color='D1D5DB')
    return Border(top=side, bottom=side, left=side, right=side)


def write_headers(ws, headers: List[str], branding: Dict, row: int = 1):
    """Write header row with styling."""
    hf = header_fill(branding)
    hfont = header_font()
    border = thin_border()
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = hfont
        cell.fill = hf
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border


def write_data_row(ws, values: List[Any], row: int, is_alt: bool = False):
    """Write a data row with optional alternating background."""
    dfont = data_font()
    border = thin_border()
    af = alt_fill() if is_alt else PatternFill()
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = dfont
        cell.border = border
        if is_alt:
            cell.fill = af
        # Right-align numbers
        if isinstance(val, (int, float)):
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = '#,##0' if isinstance(val, int) else '#,##0.00'


def auto_size_columns(ws, min_width: int = 10, max_width: int = 40):
    """Auto-size columns based on content."""
    for col in range(1, ws.max_column + 1):
        max_len = min_width
        col_letter = get_column_letter(col)
        for row in range(1, min(ws.max_row + 1, 500)):  # Sample first 500 rows
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[col_letter].width = min(max_len, max_width)


# ── Tabs ─────────────────────────────────────────────────────────────────────

def build_summary_tab(wb: Workbook, payload: Dict, agg: Dict, branding: Dict):
    """Tab 1: Summary — key metrics as a formatted summary sheet."""
    ws = wb.active
    ws.title = 'Summary'

    series = payload.get('series', {})
    channels = payload.get('channels', [])
    days = payload.get('broadcastDays', [])

    # Title section
    ws.merge_cells('A1:D1')
    cell = ws.cell(row=1, column=1, value=f"Viewership Report — {series.get('name', 'Unknown')}")
    cell.font = Font(name='Arial', size=14, bold=True, color='374151')
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:D2')
    scope_label = build_scope_label(payload)
    cell = ws.cell(row=2, column=1, value=scope_label)
    cell.font = Font(name='Arial', size=11, color='9CA3AF')
    cell.alignment = Alignment(horizontal='center')

    # Key metrics table
    write_headers(ws, ['Metric', 'Value'], branding, row=4)

    metrics_data = [
        ('Peak CCV', agg['peakCCV']),
        ('Peak Timestamp', format_datetime(agg.get('peakTimestamp'))),
        ('Average CCV', agg['avgCCV']),
        ('Total Viewed Hours', round(agg['totalViewedHours'], 2)),
        ('Tracked Channels', len(channels)),
        ('Broadcast Days', len(days)),
        ('Total Snapshots', payload.get('snapshotCount', 0)),
        ('Series', series.get('name', '')),
        ('Game', series.get('game', '')),
        ('Partner', series.get('partner', '')),
        ('Start Date', format_date(series.get('startDate'))),
        ('End Date', format_date(series.get('endDate'))),
    ]

    for i, (metric, val) in enumerate(metrics_data):
        write_data_row(ws, [metric, val], row=5 + i, is_alt=i % 2 == 1)

    # Platform breakdown section
    plats = agg.get('platformBreakdown', [])
    if plats:
        start_row = 5 + len(metrics_data) + 2
        ws.cell(row=start_row - 1, column=1, value='Platform Breakdown').font = Font(
            name='Arial', size=12, bold=True, color='374151')
        write_headers(ws, ['Platform', 'Total CCV', 'Avg CCV', 'Peak CCV'], branding, row=start_row)
        for i, p in enumerate(plats):
            write_data_row(ws, [
                platform_label(p.get('platform', '')),
                p.get('totalCCV', 0),
                round(p.get('avgCCV', 0)),
                p.get('peakCCV', 0),
            ], row=start_row + 1 + i, is_alt=i % 2 == 1)

    auto_size_columns(ws)


def build_channels_tab(wb: Workbook, payload: Dict, agg: Dict, branding: Dict):
    """Tab 2: Channels — per-channel breakdown."""
    ws = wb.create_sheet('Channels')

    headers = ['#', 'Channel', 'Platform', 'Language', 'Region', 'Category',
               'Source', 'Peak CCV', 'Avg CCV']
    write_headers(ws, headers, branding)

    leaderboard = {ch.get('channelId', ''): ch for ch in agg.get('channelLeaderboard', [])}
    channels = payload.get('channels', [])

    for i, ch in enumerate(channels):
        cid = ch.get('id', '')
        lb = leaderboard.get(cid, {})
        write_data_row(ws, [
            i + 1,
            ch.get('displayName', ''),
            platform_label(ch.get('platform', '')),
            (ch.get('language') or '—').upper(),
            ch.get('region') or '—',
            tier_label(ch.get('tier', '')),
            ch.get('source', ''),
            lb.get('peakCCV', 0),
            round(lb.get('avgCCV', 0)),
        ], row=2 + i, is_alt=i % 2 == 1)

    auto_size_columns(ws)


def build_time_series_tab(wb: Workbook, time_series: List[Dict], branding: Dict):
    """Tab 3: Time Series — raw time-bucketed data."""
    ws = wb.create_sheet('Time Series')

    headers = ['Timestamp', 'Total CCV', 'Channel Count']
    write_headers(ws, headers, branding)

    for i, dp in enumerate(time_series):
        write_data_row(ws, [
            dp.get('timestamp', ''),
            dp.get('totalCCV', 0),
            dp.get('channelCount', 0),
        ], row=2 + i, is_alt=i % 2 == 1)

    auto_size_columns(ws)


def build_platform_split_tab(wb: Workbook, payload: Dict, branding: Dict):
    """Tab 4: Platform Split — platform breakdown per broadcast day."""
    ws = wb.create_sheet('Platform Split')

    # Collect all platforms
    all_platforms = set()
    for m in payload.get('metrics', []):
        for p in m.get('platformBreakdown', []):
            all_platforms.add(p.get('platform', 'unknown'))
    plat_list = sorted(all_platforms)

    if not plat_list:
        ws.cell(row=1, column=1, value='No platform data available')
        return

    # Headers: Day | Platform1 Total | Platform1 Peak | ...
    headers = ['Broadcast Day', 'Date']
    for plat in plat_list:
        label = platform_label(plat)
        headers.extend([f'{label} Total CCV', f'{label} Peak CCV'])
    write_headers(ws, headers, branding)

    days = {d['id']: d for d in payload.get('broadcastDays', [])}

    for i, m in enumerate(payload.get('metrics', [])):
        day_id = m.get('broadcastDayId', '')
        day = days.get(day_id, {})
        row_data = [day.get('label', ''), format_date(day.get('date'))]

        plat_map = {p.get('platform', ''): p for p in m.get('platformBreakdown', [])}
        for plat in plat_list:
            pd = plat_map.get(plat, {})
            row_data.append(pd.get('totalCCV', 0))
            row_data.append(pd.get('peakCCV', 0))

        write_data_row(ws, row_data, row=2 + i, is_alt=i % 2 == 1)

    auto_size_columns(ws)


def build_language_split_tab(wb: Workbook, payload: Dict, branding: Dict):
    """Tab 5: Language Split — language breakdown per broadcast day."""
    ws = wb.create_sheet('Language Split')

    all_langs = set()
    for m in payload.get('metrics', []):
        for l in m.get('languageBreakdown', []):
            all_langs.add(l.get('language', 'unknown'))
    lang_list = sorted(all_langs)

    if not lang_list:
        ws.cell(row=1, column=1, value='No language data available')
        return

    headers = ['Broadcast Day', 'Date']
    for lang in lang_list:
        headers.extend([f'{lang.upper()} Total CCV', f'{lang.upper()} Peak CCV'])
    write_headers(ws, headers, branding)

    days = {d['id']: d for d in payload.get('broadcastDays', [])}

    for i, m in enumerate(payload.get('metrics', [])):
        day_id = m.get('broadcastDayId', '')
        day = days.get(day_id, {})
        row_data = [day.get('label', ''), format_date(day.get('date'))]

        lang_map = {l.get('language', ''): l for l in m.get('languageBreakdown', [])}
        for lang in lang_list:
            ld = lang_map.get(lang, {})
            row_data.append(ld.get('totalCCV', 0))
            row_data.append(ld.get('peakCCV', 0))

        write_data_row(ws, row_data, row=2 + i, is_alt=i % 2 == 1)

    auto_size_columns(ws)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    input_data = read_stdin_json()
    payload = input_data['payload']
    time_series = input_data.get('timeSeries', [])
    branding = input_data.get('branding', load_branding())
    output_path = input_data['outputPath']

    agg = aggregate_metrics(payload.get('metrics', []))

    wb = Workbook()

    build_summary_tab(wb, payload, agg, branding)
    build_channels_tab(wb, payload, agg, branding)
    build_time_series_tab(wb, time_series, branding)
    build_platform_split_tab(wb, payload, branding)
    build_language_split_tab(wb, payload, branding)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    print(json.dumps({'status': 'ok', 'path': output_path}))


if __name__ == '__main__':
    main()
